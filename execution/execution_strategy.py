# Copyright (C) 2026 Aguilar. This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by the Free Software Foundation,
# either version 3 of the License, or any later version.
"""Execution Strategy Engine: autonomous decision-making for the agent runtime.

Responsibilities:
1. Decide whether a task should be processed internally (artifact/sandbox) vs.
   using the user's desktop (native tools, UI automation, Office COM).
2. Generate script-based recovery approaches when a primary tool path fails.
3. Detect tool gaps and propose sandbox/dynamic_tool alternatives.
4. Build multi-tool orchestration scripts that chain operations the agent would
   otherwise need multiple separate tool calls for.
"""
from __future__ import annotations

import re
import textwrap
from typing import Any, Dict, List, Optional, Tuple

INTERNAL_PROCESSING_TOOLS = {"artifact", "sandbox", "memory", "dynamic_tool"}
DESKTOP_TOOLS = {"desktop", "windows_ui", "office", "browser"}
DISCOVERY_TOOLS = {"share_discovery", "document_intelligence", "software_inventory", "driver_manager"}
IO_TOOLS = {"filesystem", "shell", "web", "package_manager", "env_manager", "os"}

FILE_EXTENSIONS_INTERNAL = {
    ".csv", ".json", ".txt", ".log", ".xml", ".yaml", ".yml",
    ".md", ".tsv", ".ini", ".cfg", ".conf", ".html", ".htm",
}
FILE_EXTENSIONS_ARTIFACT = {
    ".pdf", ".docx", ".xlsx", ".xls", ".doc", ".msg", ".eml",
}
FILE_EXTENSIONS_DESKTOP = {
    ".pptx", ".ppt", ".accdb", ".mdb", ".vsd", ".vsdx",
    ".one", ".pub",
}

SCRIPT_TEMPLATES = {
    "read_excel_data": textwrap.dedent("""\
        import json
        try:
            import openpyxl
            wb = openpyxl.load_workbook("{path}", read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            wb.close()
            print(json.dumps({{"sheets": wb.sheetnames if hasattr(wb, 'sheetnames') else [], "rows": rows[:200], "total_rows": len(rows)}}, ensure_ascii=False, indent=2))
        except ImportError:
            import subprocess, sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
            import openpyxl
            wb = openpyxl.load_workbook("{path}", read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            wb.close()
            print(json.dumps({{"rows": rows[:200], "total_rows": len(rows)}}, ensure_ascii=False, indent=2))
    """),

    "cross_reference_report": textwrap.dedent("""\
        import json, csv, io
        data_a = {data_a_json}
        data_b = {data_b_json}
        # Cross-reference logic
        key_field_a = "{key_a}"
        key_field_b = "{key_b}"
        index_b = {{row.get(key_field_b, ""): row for row in data_b if isinstance(row, dict)}}
        results = []
        for row_a in data_a:
            if not isinstance(row_a, dict):
                continue
            key = row_a.get(key_field_a, "")
            match = index_b.get(key)
            results.append({{**row_a, "matched": match is not None, "match_data": match}})
        print(json.dumps({{"total": len(results), "matched": sum(1 for r in results if r["matched"]), "results": results[:100]}}, ensure_ascii=False, indent=2))
    """),

    "generate_report": textwrap.dedent("""\
        import json, datetime
        data = {data_json}
        report_lines = []
        report_lines.append(f"# Relatorio gerado em {{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}}")
        report_lines.append(f"")
        report_lines.append(f"Total de registros: {{len(data)}}")
        report_lines.append(f"")
        for i, item in enumerate(data[:50], 1):
            if isinstance(item, dict):
                line = " | ".join(f"{{k}}: {{v}}" for k, v in item.items())
                report_lines.append(f"{{i}}. {{line}}")
            else:
                report_lines.append(f"{{i}}. {{item}}")
        report = "\\n".join(report_lines)
        output_path = r"{output_path}"
        if output_path:
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(report)
            print(json.dumps({{"status": "saved", "path": output_path, "lines": len(report_lines)}}, ensure_ascii=False))
        else:
            print(report)
    """),

    "fallback_file_operation": textwrap.dedent("""\
        import os, json, shutil
        path = r"{path}"
        operation = "{operation}"
        if operation == "read":
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
            print(json.dumps({{"path": path, "size": len(content), "content": content[:10000]}}, ensure_ascii=False))
        elif operation == "list":
            entries = []
            for entry in os.scandir(path):
                entries.append({{"name": entry.name, "is_dir": entry.is_dir(), "size": entry.stat().st_size if not entry.is_dir() else 0}})
            print(json.dumps({{"path": path, "entries": entries}}, ensure_ascii=False, indent=2))
        elif operation == "stat":
            st = os.stat(path)
            print(json.dumps({{"path": path, "size": st.st_size, "modified": st.st_mtime, "exists": True}}, ensure_ascii=False))
    """),
}


class ExecutionStrategy:
    """Decides how to execute tasks and generates alternative approaches."""

    def decide_execution_mode(
        self,
        *,
        tool: str,
        action: str,
        params: Dict[str, Any],
        available_tools: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """Decide if a task should run internally, on desktop, or via script.

        Returns a dict with:
        - mode: "internal" | "desktop" | "script" | "native"
        - reason: human-readable explanation
        - alternative: optional dict with tool/action/params for script approach
        """
        if tool in INTERNAL_PROCESSING_TOOLS:
            return {"mode": "internal", "reason": f"{tool} is already an internal processing tool", "alternative": None}

        if tool in DESKTOP_TOOLS:
            internal_alt = self._check_desktop_internalization(tool, action, params)
            if internal_alt:
                return {
                    "mode": "internal",
                    "reason": internal_alt["reason"],
                    "alternative": internal_alt,
                }
            return {"mode": "desktop", "reason": f"{tool}.{action} requires desktop interaction", "alternative": None}

        path = params.get("path") or ""
        if path:
            ext = self._file_extension(path)
            if ext in FILE_EXTENSIONS_INTERNAL:
                return {
                    "mode": "internal",
                    "reason": f"File type {ext} can be processed internally via artifact or sandbox",
                    "alternative": {"tool": "artifact", "action": "load", "params": {"path": path}},
                }
            if ext in FILE_EXTENSIONS_ARTIFACT:
                return {
                    "mode": "internal",
                    "reason": f"File type {ext} supported by artifact processor",
                    "alternative": {"tool": "artifact", "action": "load", "params": {"path": path}},
                }
            if ext in FILE_EXTENSIONS_DESKTOP:
                return {"mode": "desktop", "reason": f"File type {ext} requires native desktop app", "alternative": None}

        if available_tools and tool not in (available_tools or []):
            script = self._generate_tool_gap_script(tool, action, params)
            if script:
                return {
                    "mode": "script",
                    "reason": f"Tool '{tool}' not available; generated sandbox script alternative",
                    "alternative": script,
                }

        return {"mode": "native", "reason": f"{tool}.{action} will use the native tool path", "alternative": None}

    def suggest_script_recovery(
        self,
        *,
        task: Dict[str, Any],
        error: str,
        attempt: int,
    ) -> Optional[Dict[str, Any]]:
        """When a task fails, suggest a sandbox script to achieve the same goal."""
        tool = task.get("tool", "")
        action = task.get("action", "")
        params = task.get("params") or {}
        error_lower = (error or "").lower()

        if "approval rejected" in error_lower:
            return None
        if "blocked by policy" in error_lower:
            return None

        script = self._recovery_script_for_failure(tool, action, params, error)
        if script:
            return {
                "classification": "script_recovery",
                "retryable": False,
                "needs_user": False,
                "suggested_action": "execute_script",
                "reason": f"Primary path ({tool}.{action}) failed; generated a sandbox script alternative",
                "script": script,
            }
        return None

    def generate_orchestration_script(
        self,
        *,
        goal: str,
        data_sources: List[Dict[str, Any]],
        output_format: str = "report",
        output_path: Optional[str] = None,
    ) -> Optional[Dict[str, Any]]:
        """Generate a script that orchestrates a multi-source data task.

        Used for complex requests like 'read emails, cross with spreadsheet, create report'.
        """
        if not data_sources:
            return None

        script_parts = [
            "import json, os, sys, datetime",
            "",
            "# Data collection phase",
            "collected = {}",
        ]

        for i, source in enumerate(data_sources):
            source_type = source.get("type", "file")
            source_path = source.get("path", "")
            source_key = source.get("key", f"source_{i}")

            if source_type == "outlook_emails":
                count = source.get("count", 20)
                script_parts.append(self._outlook_reader_snippet(source_key, count))
            elif source_type == "excel":
                script_parts.append(self._excel_reader_snippet(source_key, source_path))
            elif source_type == "csv":
                script_parts.append(self._csv_reader_snippet(source_key, source_path))
            elif source_type == "json":
                script_parts.append(self._json_reader_snippet(source_key, source_path))
            elif source_type == "text":
                script_parts.append(self._text_reader_snippet(source_key, source_path))
            else:
                script_parts.append(f'collected["{source_key}"] = {{"type": "{source_type}", "path": r"{source_path}", "error": "unsupported source type"}}')

        script_parts.append("")
        script_parts.append("# Processing phase")

        if len(data_sources) >= 2:
            script_parts.append(self._cross_reference_snippet(data_sources))

        script_parts.append("")
        script_parts.append("# Output phase")

        if output_format == "report":
            script_parts.append(self._report_output_snippet(output_path))
        elif output_format == "json":
            script_parts.append(self._json_output_snippet(output_path))
        elif output_format == "csv":
            script_parts.append(self._csv_output_snippet(output_path))
        else:
            script_parts.append(self._report_output_snippet(output_path))

        code = "\n".join(script_parts)
        return {
            "tool": "sandbox",
            "action": "execute_python",
            "params": {
                "code": code,
                "timeout": 120,
            },
        }

    def detect_tool_gap(
        self,
        *,
        tool: str,
        action: str,
        available_tools: List[str],
    ) -> Optional[Dict[str, Any]]:
        """Detect when a requested tool/action doesn't exist and propose alternatives."""
        if tool in available_tools:
            return None

        suggestion = self._generate_tool_gap_script(tool, action, {})
        if suggestion:
            return {
                "gap_detected": True,
                "missing_tool": tool,
                "missing_action": action,
                "suggestion": suggestion,
                "reason": f"Tool '{tool}' is not available; a sandbox script can achieve the same goal",
            }

        return {
            "gap_detected": True,
            "missing_tool": tool,
            "missing_action": action,
            "suggestion": None,
            "reason": f"Tool '{tool}' is not available and no automatic script alternative was found",
        }

    def _check_desktop_internalization(self, tool: str, action: str, params: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Check if a desktop tool action can be handled internally."""
        if tool == "office" and action == "excel_read_range":
            path = params.get("path")
            if path:
                return {
                    "tool": "sandbox",
                    "action": "execute_python",
                    "params": {
                        "code": SCRIPT_TEMPLATES["read_excel_data"].format(path=path),
                        "timeout": 30,
                    },
                    "reason": "Excel data can be read internally via openpyxl without opening the desktop app",
                }

        if tool == "office" and action == "outlook_search_messages":
            return None  # outlook_search_messages already works headlessly via COM

        if tool == "office" and action in {"word_find_text", "inspect_document"}:
            path = params.get("path", "")
            ext = self._file_extension(path)
            if ext in {".docx", ".pdf", ".txt"}:
                return {
                    "tool": "artifact",
                    "action": "load",
                    "params": {"path": path},
                    "reason": f"Document ({ext}) can be loaded and searched internally without desktop app",
                }

        return None

    def _recovery_script_for_failure(self, tool: str, action: str, params: Dict[str, Any], error: str) -> Optional[Dict[str, Any]]:
        error_lower = error.lower()

        if tool == "office" and "com" in error_lower:
            if action in {"excel_read_range", "list_workbook_sheets"}:
                path = params.get("path", "")
                if path:
                    return {
                        "tool": "sandbox",
                        "action": "execute_python",
                        "params": {
                            "code": SCRIPT_TEMPLATES["read_excel_data"].format(path=path),
                            "timeout": 30,
                        },
                    }
            if action in {"word_find_text", "inspect_document", "open_document"}:
                path = params.get("path", "")
                if path:
                    return {
                        "tool": "artifact",
                        "action": "load",
                        "params": {"path": path},
                    }
            if action in {"outlook_search_messages", "outlook_read_latest"}:
                return None  # COM-based office tool already works headlessly

        if tool == "filesystem" and action in {"read", "list", "stat"}:
            path = params.get("path", "")
            if path:
                return {
                    "tool": "sandbox",
                    "action": "execute_python",
                    "params": {
                        "code": SCRIPT_TEMPLATES["fallback_file_operation"].format(path=path, operation=action),
                        "timeout": 15,
                    },
                }

        if tool == "browser" and ("playwright" in error_lower or "browser" in error_lower):
            url = params.get("url", "")
            if url:
                return {
                    "tool": "sandbox",
                    "action": "execute_python",
                    "params": {
                        "code": textwrap.dedent(f"""\
                            import urllib.request, json
                            req = urllib.request.Request("{url}", headers={{"User-Agent": "Mozilla/5.0"}})
                            with urllib.request.urlopen(req, timeout=15) as resp:
                                content = resp.read().decode("utf-8", errors="replace")
                            print(json.dumps({{"url": "{url}", "status": resp.status, "content_length": len(content), "content": content[:8000]}}, ensure_ascii=False))
                        """),
                        "timeout": 20,
                    },
                }

        if tool == "document_intelligence" and action == "extract_text":
            path = params.get("path", "")
            if path:
                return {
                    "tool": "artifact",
                    "action": "load",
                    "params": {"path": path},
                }

        if tool == "share_discovery" and ("access" in error_lower or "network" in error_lower):
            path = params.get("path", "")
            if path:
                return {
                    "tool": "sandbox",
                    "action": "execute_powershell",
                    "params": {
                        "code": f'Get-ChildItem -Path "{path}" -ErrorAction SilentlyContinue | Select-Object Name, Length, LastWriteTime | ConvertTo-Json',
                        "timeout": 15,
                    },
                }

        if tool == "desktop" and action == "open_app":
            app_name = params.get("app_name", "") or params.get("app_path", "")
            if app_name:
                return {
                    "tool": "sandbox",
                    "action": "execute_powershell",
                    "params": {
                        "code": f'Start-Process "{app_name}" -PassThru | Select-Object Id, ProcessName, MainWindowTitle | ConvertTo-Json',
                        "timeout": 15,
                    },
                }

        if tool == "package_manager":
            package = params.get("package", "")
            if package and action == "install":
                return {
                    "tool": "sandbox",
                    "action": "execute_powershell",
                    "params": {
                        "code": f'winget install --id "{package}" --accept-package-agreements --accept-source-agreements 2>&1 | Out-String',
                        "timeout": 120,
                    },
                }

        if tool == "web" and action == "search_web":
            query = params.get("query", "")
            if query:
                return {
                    "tool": "sandbox",
                    "action": "execute_python",
                    "params": {
                        "code": textwrap.dedent(f"""\
                            import urllib.request, urllib.parse, json, re
                            query = "{query}"
                            url = "https://html.duckduckgo.com/html/?q=" + urllib.parse.quote(query)
                            req = urllib.request.Request(url, headers={{"User-Agent": "Mozilla/5.0"}})
                            with urllib.request.urlopen(req, timeout=10) as resp:
                                html = resp.read().decode("utf-8", errors="replace")
                            links = re.findall(r'class="result__a"[^>]*href="([^"]+)"[^>]*>([^<]+)', html)
                            results = [{{"url": u, "title": t.strip()}} for u, t in links[:10]]
                            print(json.dumps({{"query": query, "results": results}}, ensure_ascii=False, indent=2))
                        """),
                        "timeout": 15,
                    },
                }

        return None

    def _generate_tool_gap_script(self, tool: str, action: str, params: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        known_gaps = {
            ("email", "read"): lambda p: {
                "tool": "office",
                "action": "outlook_read_latest",
                "params": {"limit": p.get("count", 10)},
            },
            ("email", "search"): lambda p: {
                "tool": "office",
                "action": "outlook_search_messages",
                "params": {"query": p.get("query", ""), "limit": p.get("count", 25)},
            },
            ("report", "generate"): lambda p: {
                "tool": "sandbox",
                "action": "execute_python",
                "params": {"code": SCRIPT_TEMPLATES["generate_report"].format(
                    data_json="[]", output_path=p.get("output_path", "")
                ), "timeout": 30},
            },
            ("data", "cross_reference"): lambda p: {
                "tool": "sandbox",
                "action": "execute_python",
                "params": {"code": SCRIPT_TEMPLATES["cross_reference_report"].format(
                    data_a_json="[]", data_b_json="[]",
                    key_a=p.get("key_a", "id"), key_b=p.get("key_b", "id"),
                ), "timeout": 30},
            },
        }

        generator = known_gaps.get((tool, action))
        if generator:
            return generator(params)

        return None

    def _file_extension(self, path: str) -> str:
        if "." not in path:
            return ""
        return "." + path.rsplit(".", 1)[-1].lower()

    def _outlook_reader_snippet(self, key: str, count: int) -> str:
        return textwrap.dedent(f"""\
            try:
                import win32com.client
                outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
                inbox = outlook.GetDefaultFolder(6)
                messages = inbox.Items
                messages.Sort("[ReceivedTime]", True)
                emails = []
                for i in range(min({count}, messages.Count)):
                    msg = messages[i + 1]
                    emails.append({{"subject": msg.Subject, "sender": msg.SenderName, "received": str(msg.ReceivedTime), "body": msg.Body[:500]}})
                collected["{key}"] = emails
            except Exception as e:
                collected["{key}"] = {{"error": str(e)}}
        """)

    def _excel_reader_snippet(self, key: str, path: str) -> str:
        return textwrap.dedent(f"""\
            try:
                import openpyxl
                wb = openpyxl.load_workbook(r"{path}", read_only=True, data_only=True)
                ws = wb.active
                rows = []
                header = None
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    cells = [str(c) if c is not None else "" for c in row]
                    if i == 0:
                        header = cells
                    else:
                        if header:
                            rows.append(dict(zip(header, cells)))
                        else:
                            rows.append(cells)
                wb.close()
                collected["{key}"] = rows
            except Exception as e:
                collected["{key}"] = {{"error": str(e)}}
        """)

    def _csv_reader_snippet(self, key: str, path: str) -> str:
        return textwrap.dedent(f"""\
            try:
                import csv
                with open(r"{path}", "r", encoding="utf-8-sig", errors="replace") as f:
                    reader = csv.DictReader(f)
                    collected["{key}"] = [row for row in reader]
            except Exception as e:
                collected["{key}"] = {{"error": str(e)}}
        """)

    def _json_reader_snippet(self, key: str, path: str) -> str:
        return textwrap.dedent(f"""\
            try:
                with open(r"{path}", "r", encoding="utf-8") as f:
                    collected["{key}"] = json.load(f)
            except Exception as e:
                collected["{key}"] = {{"error": str(e)}}
        """)

    def _text_reader_snippet(self, key: str, path: str) -> str:
        return textwrap.dedent(f"""\
            try:
                with open(r"{path}", "r", encoding="utf-8", errors="replace") as f:
                    collected["{key}"] = f.read()
            except Exception as e:
                collected["{key}"] = {{"error": str(e)}}
        """)

    def _cross_reference_snippet(self, data_sources: List[Dict[str, Any]]) -> str:
        keys = [s.get("key", f"source_{i}") for i, s in enumerate(data_sources)]
        return textwrap.dedent(f"""\
            # Cross-reference available data
            all_keys = {keys}
            cross_ref = {{}}
            for k in all_keys:
                src = collected.get(k, [])
                if isinstance(src, list):
                    cross_ref[k] = {{"count": len(src), "sample": src[:3]}}
                elif isinstance(src, dict) and "error" not in src:
                    cross_ref[k] = {{"count": 1, "type": "dict"}}
                else:
                    cross_ref[k] = {{"count": 0, "error": src.get("error") if isinstance(src, dict) else "no data"}}
            collected["_cross_reference"] = cross_ref
        """)

    def _report_output_snippet(self, output_path: Optional[str]) -> str:
        save_part = ""
        if output_path:
            save_part = textwrap.dedent(f"""\
                output_path = r"{output_path}"
                with open(output_path, "w", encoding="utf-8") as f:
                    f.write(report_text)
                result["saved_to"] = output_path
            """)
        return textwrap.dedent(f"""\
            report_lines = []
            report_lines.append(f"# Relatorio - {{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}}")
            report_lines.append("")
            for source_name, source_data in collected.items():
                if source_name.startswith("_"):
                    continue
                report_lines.append(f"## {{source_name}}")
                if isinstance(source_data, list):
                    report_lines.append(f"Total registros: {{len(source_data)}}")
                    for i, item in enumerate(source_data[:30], 1):
                        if isinstance(item, dict):
                            line = " | ".join(f"{{k}}: {{v}}" for k, v in list(item.items())[:5])
                            report_lines.append(f"  {{i}}. {{line}}")
                        else:
                            report_lines.append(f"  {{i}}. {{item}}")
                elif isinstance(source_data, dict) and "error" in source_data:
                    report_lines.append(f"  ERRO: {{source_data['error']}}")
                else:
                    report_lines.append(f"  {{str(source_data)[:500]}}")
                report_lines.append("")
            report_text = "\\n".join(report_lines)
            result = {{"report": report_text, "sources": len(collected), "lines": len(report_lines)}}
            {save_part}
            print(json.dumps(result, ensure_ascii=False, indent=2))
        """)

    def _json_output_snippet(self, output_path: Optional[str]) -> str:
        save_part = ""
        if output_path:
            save_part = f"""
            with open(r"{output_path}", "w", encoding="utf-8") as f:
                json.dump(collected, f, ensure_ascii=False, indent=2, default=str)
            """
        return textwrap.dedent(f"""\
            {save_part}
            print(json.dumps(collected, ensure_ascii=False, indent=2, default=str))
        """)

    def _csv_output_snippet(self, output_path: Optional[str]) -> str:
        return textwrap.dedent(f"""\
            import csv, io
            all_rows = []
            for source_data in collected.values():
                if isinstance(source_data, list):
                    all_rows.extend(source_data)
            if all_rows and isinstance(all_rows[0], dict):
                fieldnames = list(all_rows[0].keys())
                output = io.StringIO()
                writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(all_rows)
                csv_text = output.getvalue()
                {"" if not output_path else f'with open(r"{output_path}", "w", encoding="utf-8", newline="") as f: f.write(csv_text)'}
                print(csv_text[:5000])
            else:
                print(json.dumps({{"error": "no structured data to export"}}, ensure_ascii=False))
        """)
